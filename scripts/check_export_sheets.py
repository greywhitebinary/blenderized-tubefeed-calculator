"""AppTest for the Excel export's per-blend sheet naming.

WHY THIS EXISTS. app/streamlit_app.py builds one export sheet per blend
as f"BTF {sanitize_filename(blend['name'], ...)}"[:31] with no dedupe.
openpyxl accepts a duplicate sheet title silently, and the second
to_excel() call just overwrites the first sheet in place -- so two blends
that land on the same 31-char title collapse into ONE sheet holding only
the LAST blend's ingredients. The other blend's ingredients vanish from
the export with no error, no warning, nothing on screen.

This is reachable without deliberately renaming anything: _new_blend()
re-issues "Blend 3" once the middle of three blends is deleted, and two
differently-named blends can truncate to the same 31 characters. This
test forces the collision directly (renaming two blends to the identical
name) because that's the simplest reliable repro, but the bug is the
missing dedupe, not the specific name collision.

Drives the real add-a-food (custom food from label) flow to give each
blend a distinctly-named ingredient, then intercepts the actual bytes
st.download_button() would send to the browser (AppTest has no built-in
accessor for download_button data, so this patches
streamlit.elements.widgets.button.marshall_file -- the one function both
the "Export to Excel" and "Save this day" buttons funnel through -- to
also stash (file_name, data) as it passes through on its way to the real
implementation) and reads it back with openpyxl, exactly as a real
export file opened in Excel would be.
"""

import io
import sys
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
import streamlit.elements.widgets.button as _btn_mod  # noqa: E402
from streamlit.testing.v1 import AppTest  # noqa: E402

at = AppTest.from_file(str(ROOT / "app" / "streamlit_app.py"), default_timeout=180)
at.run()
assert not at.exception, at.exception

# --- Second blend, auto-selected by _new_blend() -----------------------
next(b for b in at.button if b.label == "➕ New blend").click().run()
assert not at.exception, at.exception
bid1 = at.session_state["selected_blend_id"]
assert bid1 != 0, "New blend should not reuse the starter blend's id"


def _add_custom_ingredient(prefix: str, food_name: str) -> None:
    """Drive the add-food component's "Nutrition Facts label" mode to add
    one named custom food to whichever blend `prefix` (f"blend_{bid}")
    currently points at. Mirrors scripts/check_label_photo_fill.py."""
    mode = next(r for r in at.radio if r.key == f"{prefix}_add_mode")
    label_option = next(o for o in mode.options if "abel" in o)
    mode.set_value(label_option).run()
    next(x for x in at.text_input if x.key == f"{prefix}_cname").set_value(food_name).run()
    next(b for b in at.button if b.key == f"{prefix}_add_custom_btn").click().run()
    assert not at.exception, at.exception


_add_custom_ingredient(f"blend_{bid1}", "Alpha custom food")
next(t for t in at.text_input if t.key == f"blend_name_{bid1}").set_value("Duplicate Blend").run()
assert not at.exception, at.exception

# --- Switch back to the starter blend (id 0) via the real selector -----
# The selectbox's options are positional indices into
# list(session_state.blends.keys()), not blend ids themselves (see
# app/streamlit_app.py's "Select blend" selectbox) -- format_func maps
# index -> display name. Blend id 0 is still first in insertion order.
blend_ids_now = list(at.session_state["blends"].keys())
idx0 = blend_ids_now.index(0)
next(s for s in at.selectbox if s.key == "blend_selector").set_value(idx0).run()
assert not at.exception, at.exception
bid0 = at.session_state["selected_blend_id"]
assert bid0 == 0, bid0

_add_custom_ingredient(f"blend_{bid0}", "Beta custom food")
next(t for t in at.text_input if t.key == f"blend_name_{bid0}").set_value("Duplicate Blend").run()
assert not at.exception, at.exception

names = {bid: b["name"] for bid, b in at.session_state["blends"].items()}
assert names[bid0] == names[bid1] == "Duplicate Blend", names
print(f"OK: two blends ({bid0}, {bid1}) share the name {names[bid0]!r}, one ingredient each")

# --- Trigger the export and capture the actual bytes --------------------
_captured: list[tuple[str | None, bytes]] = []
_orig_marshall_file = _btn_mod.marshall_file


def _capturing_marshall_file(coordinates, data, proto, mimetype, file_name=None):
    _captured.append((file_name, data))
    return _orig_marshall_file(coordinates, data, proto, mimetype, file_name)


with patch.object(_btn_mod, "marshall_file", _capturing_marshall_file):
    at.run()
    assert not at.exception, at.exception

# "Export to Excel" names its file f"{sanitize_filename(recipe_name)}_report.xlsx"
# (app/streamlit_app.py); "Save this day" uses a different suggested name,
# so this filter picks out the export, not the day-save, even though both
# funnel through the same patched function on this run.
exports = [
    data for file_name, data in _captured if file_name and file_name.endswith("_report.xlsx")
]
assert exports, [fn for fn, _ in _captured]
workbook = openpyxl.load_workbook(io.BytesIO(exports[-1]))
print(f"OK: captured the real export -- {len(exports[-1])} bytes, sheets {workbook.sheetnames}")

# --- Both blends must have their OWN sheet, not one overwriting the other ---
blend_sheets = [name for name in workbook.sheetnames if name.startswith("BTF ")]
assert len(blend_sheets) == 2, (
    f"expected 2 deduped blend sheets, got {blend_sheets} -- a duplicate name "
    "collapsed two blends into one sheet"
)
assert blend_sheets[0] != blend_sheets[1], blend_sheets
print(f"OK: deduped sheet names -- {blend_sheets}")

# Each sheet's ingredient list must be the ingredient THAT blend actually
# has, not both sheets pointing at whichever blend was written last.
sheet_ingredients = {}
for sheet_name in blend_sheets:
    ws = workbook[sheet_name]
    # Row 1 is the "Measured final volume" note; row 2 is the header;
    # data starts row 3 (see the startrow=1 write in streamlit_app.py).
    sheet_ingredients[sheet_name] = ws.cell(row=3, column=1).value

food_names_found = set(sheet_ingredients.values())
assert food_names_found == {
    "Alpha custom food (custom)",
    "Beta custom food (custom)",
}, f"expected both custom foods to survive as separate sheets, got {sheet_ingredients}"
print(f"OK: both blends' ingredients survived intact -- {sheet_ingredients}")

print("\n=== EXPORT SHEET DEDUPE APPTEST PASSED ===")
